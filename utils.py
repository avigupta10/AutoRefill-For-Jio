import pandas as pd
import openpyxl
import datetime
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter


def add_sum_index(excel_name, column_name, exact_column_length, sheet_name=0):
    c = 0
    sum_index = []
    wb = openpyxl.load_workbook(excel_name)
    ws = wb.active
    df = pd.read_excel(excel_name, sheet_name=sheet_name, engine='openpyxl')
    columns = df[column_name].tolist()
    for i in range(0, len(columns) - 1):
        if columns[i + 1] != columns[i]:
            sum_index.append(int(columns.index(columns[i + 1]) + 2) + c)
            ws.insert_rows(int(columns.index(columns[i + 1]) + 2) + c)
            c += 1
    sum_index.append(exact_column_length + c + 1)
    wb.save(filename=excel_name)
    return sum_index, c


def add_sum(excel_name, sum_index, exact_column_length, total, count):
    initial_index = 2
    wb = openpyxl.load_workbook(excel_name)
    ws = wb.active
    for i in sum_index:
        print(f"worksheet['B{i}'] = '= SUM(B{initial_index}:B{i - 1})'")
        ws[f'B{i}'] = f'= SUM(C{initial_index}:B{i - 1})'
        ws[f'B{i}'].font = Font(bold=True)
        initial_index = i + 1
    ws[f'B{exact_column_length + count + 2}'] = total
    ws[f'B{exact_column_length + count + 2}'].font = Font(bold=True)
    wb.save(excel_name)
    return True


def total_sum(excel_name, column_name, sheet_name=0):
    df = pd.read_excel(excel_name, sheet_name=sheet_name, engine='openpyxl')
    columns = df[column_name].tolist()
    return sum(columns)


def time_in_range(strt, end, x):
    if strt <= end:
        return strt <= x <= end
    else:
        return strt <= x or x <= end


def get_yesterday_date_time():
    yesterday_date_time = datetime.date.today() - datetime.timedelta(days=1)
    from_ = datetime.datetime(yesterday_date_time.year, yesterday_date_time.month, yesterday_date_time.day, 12, 0)
    return from_


def get_today_date_time():
    today_date_time = datetime.date.today()
    to_ = datetime.datetime(today_date_time.year, today_date_time.month, today_date_time.day, 7, 0)
    return to_


def extract_time(excel_name, column_name, start_from=get_yesterday_date_time(), to=get_today_date_time(), sheet_name=0):
    time = []
    time_range = []
    df = pd.read_excel(excel_name, sheet_name=sheet_name, engine='openpyxl')
    columns = df[column_name].tolist()
    for i in columns:
        if str(i).split('/') == ['nan']:
            pass
        else:
            t = "".join(str(i).split('/'))
            time.append(datetime.datetime.strptime(t, '%d %b %Y %H:%M'))
    for j in time:
        if time_in_range(start_from, to, j):
            time_range.append(str(j.strftime('%#d %b %Y /%H:%M')))
    return time_range


def size_and_add_cols(file_name, colm):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    ws.insert_cols(colm)
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column name
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                max_length = 7
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    ws.column_dimensions['B'].width = 9.5
    ws['A1'] = 'Retailer'
    ws['B1'] = 'AR'
    ws['C1'] = 'Receipt'
    ws['C1'].font = Font(bold=True, size=12)
    ws['D1'] = 'Date/Time'
    ws['E1'] = 'FOS'
    cell_range = 'A1:E' + str(ws.max_row)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))  # Position of border
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border
    wb.save(file_name)
    print('added')
    return True


def fos_format(col_name):
    m = list(map(str,[" ".join(str(n).split()[1:]) for n in col_name]))
    return m