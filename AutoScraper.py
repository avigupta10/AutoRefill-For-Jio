import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from automate import *
from styleframe import StyleFrame, Styler, utils

from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
import datetime
from utils import add_sum_index, add_sum, total_sum, extract_time, size_and_add_cols, fos_format

username = ""
password = ""

# login_link = "https://partnercentral.jioconnect.com/c/portal/login?p_l_id=20187&redirect=/group/guest/home"
login_url = "https://fiori.jioconnect.com/zhttp_request"
dsm_orders_url = "https://onejio.jioconnect.com/dsm-orders/"
etopup_order_url = "https://onejio.jioconnect.com/dsm-orders/#/etopuporders"

push_button = "input[value='Push']"
filter_button = '//*[@id="root"]/div/div/div/main/div/main/div/div[1]/div[2]/div[2]/button/span[1]'
completed_button_1 = '/html/body/div[3]/div[3]/div/div/div/div[1]/div'
completed_button_2 = '//*[@id="menu-"]/div[3]/ul/li[2]'
search_button = '/html/body/div[3]/div[3]/div/div/div/div[2]/div[3]/button/span[1]'
auto_refill_button_1 = '//*[@id="root"]/div/div/div/main/div/main/div/div[2]/div/div/div[' \
                       '2]/div/div/div/table/tbody/tr[1]/td[9]/div/div/div '
auto_refill_button_2 = '//*[@id="menu-"]/div[3]/ul/li[3]/div/span'
rows_button_1 = '//*[@id="root"]/div/div/div/main/div/main/div/div[2]/div/div/table/tfoot/tr/td/div/div[2]/div/div'
rows_button_2 = '//*[@id="menu-"]/div[3]/ul/li[4]'

options = webdriver.ChromeOptions()
options.add_argument("--headless")
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_argument('window-size=1920x1080')
browser = webdriver.Chrome(options=options)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = datetime.date.today().strftime("%d %b %Y")
    file_path = f"\\{datetime.date.today().strftime('%d-%b-%Y')}.xlsx"
    file_name = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop') + file_path
    login(browser=browser, login_link=login_url, password=password, user_name=username)
    automate(
        browser, login_url, dsm_orders_url,
        etopup_order_url,
        push_button,
        filter_button,
        completed_button_1,
        completed_button_2,
        search_button,
        auto_refill_button_1,
        auto_refill_button_2,
        rows_button_1,
        rows_button_2,
        driver=browser,
    )
    r = len(browser.find_elements_by_xpath("//*[@class='MuiTable-root']/tbody/tr"))
    c = len(browser.find_elements_by_xpath("//*[@class='MuiTable-root']/tbody/tr[3]/td"))
    try:
        for row in range(2, r):
            for col in range(1, c):
                if browser.find_element_by_xpath(f"//*[@class='MuiTable-root']/tbody/tr[{row}]/td[{col}]").text:
                    data = browser.find_element_by_xpath(f"//*[@class='MuiTable-root']/tbody/tr[{row}]/td[{col}]").text
                    char = get_column_letter(col)
                    print(ws[char + str(row)])
                    ws[char + str(row)] = data
    except NoSuchElementException:
        print("End")
    wb.save(file_name)

    # Getting the exact column length for the final SUM
    exact_column_length = len(extract_time(file_name, column_name='Unnamed: 4')) + 1

    # Creating the pandas dataframe to sort and filter values
    df = pd.read_excel(file_name, engine='openpyxl', index_col=None)

    # Dropping unwanted columns and merging it
    df.drop(['Unnamed: 7'], axis=1)
    df['Unnamed: 7'] = fos_format(df.get('Unnamed: 7'))
    l1 = df['Unnamed: 1'].tolist()
    l2 = df['Unnamed: 2'].tolist()
    l3 = list(map(str, ["".join(f"{r} ({rid})") for r, rid in zip(l1, l2)]))
    df['Unnamed: 1'] = l3

    del df['Unnamed: 2']

    # Filtering the worksheet within the given time
    datetime_filter = df['Unnamed: 4'].isin(extract_time(file_name, column_name='Unnamed: 4'))
    df = df.loc[datetime_filter, ['Unnamed: 1', 'Unnamed: 3', 'Unnamed: 4', 'Unnamed: 7', ]]

    # Sorting the values
    df.sort_values(by='Unnamed: 7', ascending=True, inplace=True)

    # Styling the dataframe
    writer = StyleFrame.ExcelWriter(file_name)
    sf = StyleFrame(df, Styler(shrink_to_fit=False, wrap_text=False, horizontal_alignment='left', font_size=11))
    sf.to_excel(writer)
    writer.save()

    total = total_sum(file_name, column_name="Unnamed: 3")

    sum_index, c = add_sum_index(file_name, column_name='Unnamed: 7', exact_column_length=exact_column_length)

    add_sum(file_name, sum_index, exact_column_length, total, c)

    size_and_add_cols(file_name, 3)

    print('Final sheet saved to', file_name)
    return True


if '__main__' == __name__:
    main()
